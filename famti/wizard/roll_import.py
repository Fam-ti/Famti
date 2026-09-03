# -*- coding: utf-8 -*-

import base64
import io
import logging
from datetime import datetime, date

import openpyxl

from odoo import fields, models, _
from odoo.exceptions import UserError, ValidationError


_logger = logging.getLogger(__name__)


class RollImportWizard(models.TransientModel):
    _name = "roll.import.wizard"
    _description = "Import Rolls from Excel"

    file = fields.Binary(
        string="Excel File",
        required=True,
    )

    filename = fields.Char(
        string="Filename",
    )

    location_id = fields.Many2one(
        "stock.location",
        string="Location",
        required=True,
        default=lambda self: self.env["stock.location"].search(
            [("complete_name", "=", "FM/Stock/QC HOLD")],
            limit=1,
        ),
        readonly=True,
    )

    # ============================================================
    # MAIN IMPORT
    # ============================================================

    def action_import(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        if not self.filename:
            raise UserError(_("Please provide an Excel filename."))

        if not self.filename.lower().endswith((".xlsx", ".xlsm")):
            raise UserError(
                _("Only .xlsx or .xlsm files are supported.")
            )

        # --------------------------------------------------------
        # READ EXCEL
        # --------------------------------------------------------

        try:
            file_data = base64.b64decode(self.file)

            workbook = openpyxl.load_workbook(
                io.BytesIO(file_data),
                data_only=True,
            )

        except Exception as e:
            _logger.exception("Unable to read Excel")

            raise UserError(
                _("Unable to read Excel file:\n%s") % str(e)
            )

        if not workbook.sheetnames:
            raise UserError(
                _("The Excel file does not contain any sheet.")
            )

        sheet = workbook[workbook.sheetnames[0]]

        # --------------------------------------------------------
        # READ HEADERS
        # --------------------------------------------------------

        headers = []

        for cell in sheet[1]:
            headers.append(
                self._normalize_header(cell.value)
                if cell.value is not None
                else ""
            )

        _logger.info(
            "ROLL IMPORT HEADERS: %s",
            headers,
        )

        required_headers = [
            "product",
            "product_code",
            "roll_numbers",
        ]

        missing_headers = [
            header
            for header in required_headers
            if header not in headers
        ]

        if missing_headers:
            raise UserError(
                _(
                    "Following required Excel columns are missing:\n%s"
                )
                % "\n".join(missing_headers)
            )

        processed = 0
        created = 0
        updated = 0
        errors = []

        # ========================================================
        # PROCESS ROWS
        # ========================================================

        for excel_row_number, row in enumerate(
            sheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):

            # Ignore blank rows
            if not any(
                value is not None and str(value).strip()
                for value in row
            ):
                continue

            try:

                values = self._read_row(
                    headers,
                    row,
                )

                _logger.info(
                    "ROLL IMPORT ROW %s VALUES: %s",
                    excel_row_number,
                    values,
                )

                # ------------------------------------------------
                # BASIC
                # ------------------------------------------------

                roll_number = self._clean_string(
                    values.get("roll_numbers")
                )

                product_code = self._clean_string(
                    values.get("product_code")
                )

                if not roll_number:
                    raise ValidationError(
                        _("Roll Numbers is empty.")
                    )

                if not product_code:
                    raise ValidationError(
                        _("Product Code is empty.")
                    )

                # ------------------------------------------------
                # PRODUCT
                # ------------------------------------------------

                product = self._find_product(
                    product_code,
                    values.get("product"),
                )

                if not product:
                    raise ValidationError(
                        _(
                            "Product not found.\n"
                            "Product Code: %s\n"
                            "Product: %s"
                        )
                        % (
                            product_code,
                            values.get("product"),
                        )
                    )

                # ------------------------------------------------
                # VALIDATE NUMBERS
                # ------------------------------------------------

                self._validate_row(
                    values,
                    excel_row_number,
                )

                Lot = self.env["stock.lot"]

                # ------------------------------------------------
                # FIND EXISTING LOT
                # ------------------------------------------------

                lot = Lot.search(
                    [
                        ("name", "=", roll_number),
                        ("product_id", "=", product.id),
                    ],
                    limit=1,
                )

                # Same roll with another product
                other_lot = Lot.search(
                    [
                        ("name", "=", roll_number),
                        ("product_id", "!=", product.id),
                    ],
                    limit=1,
                )

                if other_lot:
                    raise ValidationError(
                        _(
                            "Roll Number %s already exists "
                            "against another product: %s"
                        )
                        % (
                            roll_number,
                            other_lot.product_id.display_name,
                        )
                    )

                # ------------------------------------------------
                # PREPARE ALL LOT VALUES
                # ------------------------------------------------

                lot_values = self._prepare_lot_values(
                    values,
                    product,
                )

                _logger.info(
                    "ROLL IMPORT LOT VALUES BEFORE WRITE "
                    "row=%s roll=%s: %s",
                    excel_row_number,
                    roll_number,
                    lot_values,
                )

                # ------------------------------------------------
                # CREATE / UPDATE
                # ------------------------------------------------

                if lot:

                    _logger.info(
                        "ROLL IMPORT UPDATING EXISTING LOT "
                        "id=%s name=%s",
                        lot.id,
                        lot.name,
                    )

                    lot.write(lot_values)

                    self.env.cr.flush()

                    updated += 1

                else:

                    create_values = dict(lot_values)

                    create_values.update({
                        "name": roll_number,
                        "product_id": product.id,
                    })

                    lot = Lot.create(create_values)

                    self.env.cr.flush()

                    created += 1

                # ------------------------------------------------
                # RE-BROWSE LOT FROM DATABASE
                # ------------------------------------------------

                lot = Lot.browse(lot.id)

                lot.invalidate_recordset()

                # ------------------------------------------------
                # VERIFY EVERY FIELD
                # ------------------------------------------------

                self._verify_lot(
                    lot,
                    values,
                    excel_row_number,
                )

                # ------------------------------------------------
                # EXACT QUANTITY
                # ------------------------------------------------

                quantity = self._number(
                    values.get("quantity")
                )

                if quantity is not None:

                    self._update_inventory_quantity(
                        product=product,
                        lot=lot,
                        quantity=quantity,
                        location=self.location_id,
                        row_number=excel_row_number,
                    )

                processed += 1

            except Exception as e:

                _logger.exception(
                    "ROLL IMPORT ERROR row %s",
                    excel_row_number,
                )

                errors.append(
                    _(
                        "Row %s | Roll: %s | Product: %s | %s"
                    )
                    % (
                        excel_row_number,
                        self._clean_string(
                            self._get_row_value(
                                headers,
                                row,
                                "roll_numbers",
                            )
                        ) or "None",
                        self._clean_string(
                            self._get_row_value(
                                headers,
                                row,
                                "product_code",
                            )
                        ) or "None",
                        str(e),
                    )
                )

        # ========================================================
        # RESULT
        # ========================================================

        if errors:

            message = _(
                "Import completed with errors.\n\n"
                "Processed: %s\n"
                "Created: %s\n"
                "Updated: %s\n"
                "Errors: %s\n\n"
                "%s"
            ) % (
                processed,
                created,
                updated,
                len(errors),
                "\n".join(errors[:100]),
            )

            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("Import Completed With Errors"),
                    "message": message,
                    "type": "warning",
                    "sticky": True,
                },
            }

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Roll Import Successful"),
                "message": _(
                    "Processed: %s | Created: %s | Updated: %s"
                ) % (
                    processed,
                    created,
                    updated,
                ),
                "type": "success",
                "sticky": False,
            },
        }

    # ============================================================
    # READ ROW
    # ============================================================

    def _read_row(self, headers, row):

        values = {}

        for index, header in enumerate(headers):

            if not header:
                continue

            values[header] = (
                row[index]
                if index < len(row)
                else None
            )

        return values

    def _prepare_lot_values(self, values, product):

        Lot = self.env["stock.lot"]

        vals = {}
        if "product_code" in Lot._fields:
            value = self._clean_string(
                values.get("product_code")
            )

            if value is not None:
                vals["product_code"] = value
        if "supplier_name" in Lot._fields:

            value = self._clean_string(
                values.get("supplier_name")
            )

            if value:

                supplier = self._find_partner(value)

                if not supplier:
                    raise ValidationError(
                        _("Supplier not found: %s") % value
                    )

                vals["supplier_name"] = supplier.id

        if "film" in Lot._fields:

            value = self._clean_string(
                values.get("type")
            )

            if value is not None:
                vals["film"] = value

        if "film_type" in Lot._fields:

            value = self._clean_string(
                values.get("film_type")
            )

            if value is not None:
                vals["film_type"] = value

        if "film_description" in Lot._fields:

            value = self._clean_string(
                values.get("film_description")
            )

            if value is not None:
                vals["film_description"] = value

        if "treatment_in" in Lot._fields:

            value = self._map_treatment_in(
                values.get("treatment_in")
            )

            if value is not None:
                vals["treatment_in"] = value

        if "treatment_out" in Lot._fields:

            value = self._map_treatment_out(
                values.get("treatment_out")
            )

            if value is not None:
                vals["treatment_out"] = value

        if "pallet_no" in Lot._fields:

            value = self._clean_string(
                values.get("pallet_number")
            )

            if value is not None:
                vals["pallet_no"] = value

        if "thickness" in Lot._fields:

            value = self._number(
                values.get("thickness")
            )

            if value is not None:
                vals["thickness"] = value

        if "thickness_uom" in Lot._fields:

            value = self._selection_value(
                values.get("thickness_uom"),
                {
                    "micron": "micron",
                    "microns": "micron",
                    "µm": "micron",
                    "gauge": "guage",
                    "guage": "guage",
                },
            )

            if value is not None:
                vals["thickness_uom"] = value

        if "width_val" in Lot._fields:

            value = self._number(
                values.get("width")
            )

            if value is not None:
                vals["width_val"] = value
        if "width_uom" in Lot._fields:

            value = self._selection_value(
                values.get("width_uom"),
                {
                    "mm": "mm",
                    "millimeter": "mm",
                    "millimeters": "mm",
                    "inch": "inch",
                    "inches": "inch",
                    "in": "inch",
                    '"': "inch",
                },
            )

            if value is not None:
                vals["width_uom"] = value
        if "weight" in Lot._fields:

            value = self._number(
                values.get("weight")
            )

            if value is not None:
                vals["weight"] = value

        if "weight_uom" in Lot._fields:

            value = self._selection_value(
                values.get("weight_uom"),
                {
                    "kg": "kg",
                    "kgs": "kg",
                    "kilogram": "kg",
                    "kilograms": "kg",

                    "lb": "lbs",
                    "lbs": "lbs",
                    "pound": "lbs",
                    "pounds": "lbs",

                    "g": "gm",
                    "gm": "gm",
                    "gram": "gm",
                    "grams": "gm",
                },
            )

            if value is not None:
                vals["weight_uom"] = value

        if "length_val" in Lot._fields:

            value = self._number(
                values.get("length")
            )

            if value is not None:
                vals["length_val"] = value

        if "length_uom" in Lot._fields:

            value = self._selection_value(
                values.get("length_uom"),
                {
                    "m": "m",
                    "meter": "m",
                    "meters": "m",

                    "ft": "feet",
                    "foot": "feet",
                    "feet": "feet",
                },
            )

            if value is not None:
                vals["length_uom"] = value

        if "received_date" in Lot._fields:

            value = self._parse_date(
                values.get("received_date")
            )

            if value is not None:
                vals["received_date"] = value


        if "aging" in Lot._fields:

            value = self._clean_string(
                values.get("aging")
            )

            if value is not None:
                vals["aging"] = value


        if "core_id" in Lot._fields:

            value = self._clean_string(
                values.get("core_id")
            )

            if value is not None:
                vals["core_id"] = value


        if "no_of_joint" in Lot._fields:

            value = self._clean_string(
                values.get("no_of_joint")
            )

            if value is not None:
                vals["no_of_joint"] = value


        if "lot_number" in Lot._fields:

            value = self._clean_string(
                values.get("lot_number")
            )

            if value is not None:
                vals["lot_number"] = value

        # ========================================================
        # IMPORTANT:
        # DO NOT MAP TYPE INTO CATEGORY.
        #
        # Your Excel "Type = BARE"
        # goes ONLY into:
        #
        # film = BARE
        #
        # category is untouched because Excel has no Category column.
        # ========================================================

        # ========================================================
        # DEBUG
        # ========================================================

        _logger.info(
            "FINAL LOT VALUES TO WRITE: %s",
            vals,
        )

        return vals


    def _verify_lot(
        self,
        lot,
        values,
        row_number,
    ):

        lot.invalidate_recordset()

        fields_to_check = [
            "product_code",
            "supplier_name",
            "film",
            "film_type",
            "film_description",
            "treatment_in",
            "treatment_out",
            "pallet_no",
            "thickness",
            "thickness_uom",
            "width_val",
            "width_uom",
            "weight",
            "weight_uom",
            "length_val",
            "length_uom",
            "received_date",
            "aging",
            "core_id",
            "no_of_joint",
            "lot_number",
        ]

        result = {}

        for field_name in fields_to_check:

            if field_name not in lot._fields:
                continue

            value = lot[field_name]

            if field_name == "supplier_name":
                value = (
                    value.display_name
                    if value
                    else None
                )

            result[field_name] = value

        _logger.info(
            "=================================================="
        )

        _logger.info(
            "ROLL IMPORT DATABASE VERIFICATION "
            "ROW %s LOT %s",
            row_number,
            lot.name,
        )

        _logger.info(
            "DATABASE VALUES: %s",
            result,
        )

        _logger.info(
            "TYPE / film = %s",
            lot.film if "film" in lot._fields else None,
        )

        _logger.info(
            "FILM TYPE = %s",
            lot.film_type if "film_type" in lot._fields else None,
        )

        _logger.info(
            "CORE ID = %s",
            lot.core_id if "core_id" in lot._fields else None,
        )

        _logger.info(
            "THICKNESS = %s",
            lot.thickness if "thickness" in lot._fields else None,
        )

        _logger.info(
            "WIDTH = %s",
            lot.width_val if "width_val" in lot._fields else None,
        )

        _logger.info(
            "WEIGHT = %s",
            lot.weight if "weight" in lot._fields else None,
        )

        _logger.info(
            "LENGTH = %s",
            lot.length_val if "length_val" in lot._fields else None,
        )

        _logger.info(
            "PALLET = %s",
            lot.pallet_no if "pallet_no" in lot._fields else None,
        )

        _logger.info(
            "=================================================="
        )


    def _update_inventory_quantity(
        self,
        product,
        lot,
        quantity,
        location,
        row_number,
    ):

        if quantity is None:
            return

        if not location:
            raise ValidationError(
                _("Inventory Location is required.")
            )

        Quant = self.env["stock.quant"]

        # ========================================================
        # NO CONVERSION
        #
        # Excel:
        #
        # Quantity = 483.25
        # Quantity UOM = lbs
        #
        # We DO NOT convert this to kg.
        #
        # Odoo inventory quantity receives:
        #
        # 483.25
        #
        # EXACTLY.
        # ========================================================

        exact_quantity = quantity

        _logger.info(
            "=================================================="
        )

        _logger.info(
            "QUANTITY UPDATE ROW %s",
            row_number,
        )

        _logger.info(
            "EXCEL QUANTITY = %s",
            quantity,
        )

        _logger.info(
            "QUANTITY SENT TO STOCK = %s",
            exact_quantity,
        )

        quant = Quant.with_context(
            inventory_mode=True
        ).search(
            [
                ("product_id", "=", product.id),
                ("lot_id", "=", lot.id),
                ("location_id", "=", location.id),
                ("package_id", "=", False),
                ("owner_id", "=", False),
            ],
            limit=1,
        )


        if not quant:

            _logger.info(
                "CREATING QUANT WITH EXACT QUANTITY = %s",
                exact_quantity,
            )

            quant = Quant.with_context(
                inventory_mode=True
            ).create(
                {
                    "product_id": product.id,
                    "location_id": location.id,
                    "lot_id": lot.id,
                    "inventory_quantity": exact_quantity,
                }
            )


        else:

            _logger.info(
                "EXISTING QUANT id=%s OLD quantity=%s",
                quant.id,
                quant.quantity,
            )

            quant.with_context(
                inventory_mode=True
            ).write(
                {
                    "inventory_quantity": exact_quantity,
                }
            )


        quant.with_context(
            inventory_mode=True
        ).action_apply_inventory()

        quant.invalidate_recordset()

        _logger.info(
            "AFTER APPLY: QUANT id=%s quantity=%s",
            quant.id,
            quant.quantity,
        )

        # IMPORTANT:
        # product_qty is computed by Odoo from stock.quant.
        # We NEVER write product_qty directly.

        self.env.cr.flush()

        lot.invalidate_recordset()

        _logger.info(
            "AFTER APPLY: LOT=%s product_qty=%s",
            lot.name,
            lot.product_qty,
        )

        _logger.info(
            "=================================================="
        )

    def _find_product(self, product_code, product_name=None):

        Product = self.env["product.product"]
        Template = self.env["product.template"]

        product_code = self._clean_string(product_code)
        product_name = self._clean_string(product_name)

        _logger.info(
            "========== PRODUCT SEARCH =========="
        )
        _logger.info(
            "Product Code from Excel: %r",
            product_code,
        )
        _logger.info(
            "Product Name from Excel: %r",
            product_name,
        )

        # ------------------------------------------------------------
        # 1. PRODUCT VARIANT - DEFAULT CODE
        # ------------------------------------------------------------

        if product_code:

            product = Product.search(
                [
                    ("default_code", "=", product_code),
                ],
                limit=1,
            )

            _logger.info(
                "Variant default_code search result: %s",
                product.ids,
            )

            if product:
                return product

        # ------------------------------------------------------------
        # 2. PRODUCT VARIANT - BARCODE
        # ------------------------------------------------------------

        if product_code:

            product = Product.search(
                [
                    ("barcode", "=", product_code),
                ],
                limit=1,
            )

            _logger.info(
                "Variant barcode search result: %s",
                product.ids,
            )

            if product:
                return product

        # ------------------------------------------------------------
        # 3. PRODUCT TEMPLATE - DEFAULT CODE
        # ------------------------------------------------------------

        if product_code:

            template = Template.search(
                [
                    ("default_code", "=", product_code),
                ],
                limit=1,
            )

            _logger.info(
                "Template default_code search result: %s",
                template.ids,
            )

            if template:

                product = template.product_variant_id

                if product:
                    return product

                product = template.product_variant_ids[:1]

                if product:
                    return product

        # ------------------------------------------------------------
        # 4. PRODUCT NAME
        #
        # Odoo 18 product.template.name is translated JSONB.
        # Use the ORM's name search instead of comparing name directly.
        # ------------------------------------------------------------

        if product_name:

            templates = Template.with_context(
                lang="en_US"
            ).search(
                [
                    ("name", "=", product_name),
                ],
                limit=1,
            )

            _logger.info(
                "Template name search result: %s",
                templates.ids,
            )

            if templates:

                product = templates.product_variant_id

                if product:
                    return product

                product = templates.product_variant_ids[:1]

                if product:
                    return product

        # ------------------------------------------------------------
        # 5. NAME SEARCH - CASE INSENSITIVE FALLBACK
        # ------------------------------------------------------------

        if product_name:

            templates = Template.with_context(
                lang="en_US"
            ).search(
                [
                    ("name", "ilike", product_name),
                ],
                limit=1,
            )

            _logger.info(
                "Template name ilike search result: %s",
                templates.ids,
            )

            if templates:

                product = templates.product_variant_id

                if product:
                    return product

                product = templates.product_variant_ids[:1]

                if product:
                    return product

        _logger.warning(
            "PRODUCT NOT FOUND | code=%r | name=%r",
            product_code,
            product_name,
        )

        return Product.browse()

    def _find_partner(self, value):

        Partner = self.env["res.partner"]

        value = self._clean_string(value)

        if not value:
            return False

        partner = Partner.search(
            [
                (
                    "name",
                    "=",
                    value,
                ),
            ],
            limit=1,
        )

        if partner:
            return partner

        return Partner.search(
            [
                (
                    "name",
                    "ilike",
                    value,
                ),
            ],
            limit=1,
        )


    def _validate_row(
        self,
        values,
        row_number,
    ):

        numeric_fields = [
            "thickness",
            "width",
            "weight",
            "quantity",
            "length",
        ]

        for field_name in numeric_fields:

            raw_value = values.get(
                field_name
            )

            if raw_value in (
                None,
                "",
            ):
                continue

            try:
                self._number(raw_value)

            except Exception:

                raise ValidationError(
                    _(
                        "%s is not a valid number: %s"
                    )
                    % (
                        field_name,
                        raw_value,
                    )
                )

        quantity = self._number(
            values.get("quantity")
        )

        if quantity is not None and quantity < 0:
            raise ValidationError(
                _("Quantity cannot be negative.")
            )

    def _map_treatment_in(self, value):

        value = self._clean_string(value)

        if not value:
            return None

        mapping = {

            # ------------------------------------------------------------
            # CORONA
            # ------------------------------------------------------------
            "corona": "corona",

            # ------------------------------------------------------------
            # METALLIZED ON CORONA
            # ------------------------------------------------------------
            "met on corona": "met_corona",
            "met corona": "met_corona",
            "metallized on corona": "met_corona",
            "metallised on corona": "met_corona",

            "met on chemical": "met_chemical",
            "met chemical": "met_chemical",
            "metallized on chemical": "met_chemical",
            "metallised on chemical": "met_chemical",

            "met on plain": "met_plain",
            "met plain": "met_plain",
            "metallized on plain": "met_plain",
            "metallised on plain": "met_plain",

            "met on copolymer": "met_copolymer",
            "met copolymer": "met_copolymer",
            "metallized on copolymer": "met_copolymer",
            "metallised on copolymer": "met_copolymer",
            "plain": "plain",
            "pvdc": "pvdc",
            "pvdc coated": "pvdc",
            "pvdc coated film": "pvdc",
            "soft touch": "soft_touch",
            "soft-touch": "soft_touch",
            "alox": "alox",
            "top coat alox": "alox",
            "topcoat alox": "alox",
            "top coat aloxed": "alox",
            "chemical coated": "chemical_coat",
            "chemical coat": "chemical_coat",
            "chemical coating": "chemical_coat",
            "acrylic": "acrylic",
            "acrylic coated": "acrylic",
            "copolymer": "copolymer",
            "co-polymer": "copolymer",
            "co polymer": "copolymer",
            "special chemical": "special_chemical",
            "special chemical coated": "special_chemical",
            "special chemical coating": "special_chemical",
            "sp. chemical":"special_chemical",
        }

        result = mapping.get(value.lower())

        if result is None:
            raise ValidationError(
                _("Invalid Treatment IN value: %s") % value
            )

        return result

    def _map_treatment_out(self, value):

        value = self._clean_string(value)

        if not value:
            return None

        mapping = {

            # ------------------------------------------------------------
            # CORONA
            # ------------------------------------------------------------
            "corona": "corona",

            # ------------------------------------------------------------
            # METALLIZED ON CORONA
            # ------------------------------------------------------------
            "met on corona": "met_corona",
            "met corona": "met_corona",
            "metallized on corona": "met_corona",
            "metallised on corona": "met_corona",

            # Existing client wording
            "metallized on corona outside": "met_corona_out",
            "met corona outside": "met_corona_out",
            "metallised on corona outside": "met_corona_out",

            # ------------------------------------------------------------
            # METALLIZED ON CHEMICAL
            # ------------------------------------------------------------
            "met on chemical": "met_chemical",
            "met chemical": "met_chemical",
            "metallized on chemical": "met_chemical",
            "metallised on chemical": "met_chemical",

            # ------------------------------------------------------------
            # METALLIZED ON PLAIN
            # ------------------------------------------------------------
            "met on plain": "met_plain",
            "met plain": "met_plain",
            "metallized on plain": "met_plain",
            "metallised on plain": "met_plain",

            # ------------------------------------------------------------
            # METALLIZED ON COPOLYMER
            # ------------------------------------------------------------
            "met on copolymer": "met_copolymer",
            "met copolymer": "met_copolymer",
            "metallized on copolymer": "met_copolymer",
            "metallised on copolymer": "met_copolymer",

            # ------------------------------------------------------------
            # PLAIN
            # ------------------------------------------------------------
            "plain": "plain",

            # ------------------------------------------------------------
            # PVDC
            # ------------------------------------------------------------
            "pvdc": "pvdc_out",
            "pvdc coated": "pvdc_out",
            "pvdc coated film": "pvdc_out",

            # ------------------------------------------------------------
            # ACRYLIC
            # ------------------------------------------------------------
            "acrylic": "acrylic",
            "acrylic coated": "acrylic",

            # ------------------------------------------------------------
            # COPOLYMER
            # ------------------------------------------------------------
            "copolymer": "copolymer",
            "co-polymer": "copolymer",
            "co polymer": "copolymer",

            # ------------------------------------------------------------
            # SOFT TOUCH
            # ------------------------------------------------------------
            "soft touch": "soft_touch",
            "soft-touch": "soft_touch",

            # ------------------------------------------------------------
            # ALOX
            # ------------------------------------------------------------
            "alox": "alox",
            "top coat alox": "alox",
            "topcoat alox": "alox",

            # ------------------------------------------------------------
            # CHEMICAL COATED
            # ------------------------------------------------------------
            "chemical coated": "chemical_coat",
            "chemical coat": "chemical_coat",
            "chemical coating": "chemical_coat",

            # ------------------------------------------------------------
            # SPECIAL CHEMICAL
            # ------------------------------------------------------------
            "special chemical": "special_chemical",
            "special chemical coated": "special_chemical",
            "special chemical coating": "special_chemical",
        }

        result = mapping.get(value.lower())

        if result is None:
            raise ValidationError(
                _("Invalid Treatment OUT value: %s") % value
            )

        return result

    def _selection_value(
        self,
        value,
        mapping,
    ):

        value = self._clean_string(value)

        if not value:
            return None

        key = value.lower()

        result = mapping.get(key)

        if result is None:
            raise ValidationError(
                _("Invalid value: %s") % value
            )

        return result


    def _number(self, value):

        if value is None:
            return None

        if isinstance(value, bool):
            raise ValidationError(
                _("Boolean value cannot be used as a number.")
            )

        if isinstance(value, (int, float)):
            return float(value)

        value = str(value).strip()

        if not value:
            return None

        # Only remove Excel thousands separator.
        # NO UNIT CONVERSION.
        value = value.replace(",", "")

        return float(value)

    def _parse_date(self, value):

        if value is None:
            return None

        # ------------------------------------------------------------
        # Excel datetime
        # ------------------------------------------------------------
        if isinstance(value, datetime):
            return value.date()

        # ------------------------------------------------------------
        # Excel date
        # ------------------------------------------------------------
        if isinstance(value, date):
            return value

        # ------------------------------------------------------------
        # String date
        # ------------------------------------------------------------
        value = str(value).strip()

        if not value:
            return None

        formats = [
            "%m/%d/%y",  # 2/25/25
            "%m/%d/%Y",  # 2/25/2025
            "%m-%d-%y",  # 2-25-25
            "%m-%d-%Y",  # 2-25-2025

            "%d/%m/%y",
            "%d/%m/%Y",
            "%d-%m-%y",
            "%d-%m-%Y",

            "%Y-%m-%d",
            "%Y/%m/%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        raise ValidationError(
            _("Invalid Received date: %s") % value
        )

    def _clean_string(self, value):

        if value is None:
            return None

        if isinstance(value, float):
            if value.is_integer():
                value = int(value)

        value = str(value)

        # Replace Excel/non-breaking spaces
        value = value.replace("\u00A0", " ")
        value = value.replace("\u200B", "")

        value = value.strip()

        if not value:
            return None

        return value


    def _normalize_header(self, value):

        value = str(
            value
        ).strip().lower()

        value = value.replace(
            "\n",
            " ",
        )

        value = value.replace(
            "-",
            " ",
        )

        while "  " in value:
            value = value.replace(
                "  ",
                " ",
            )

        mapping = {

            "product": "product",

            "product code": "product_code",

            "roll numbers": "roll_numbers",
            "roll number": "roll_numbers",
            "serial numbers": "roll_numbers",
            "serial number": "roll_numbers",

            "supplier name": "supplier_name",
            "supplier": "supplier_name",

            "film type": "film_type",

            "type": "type",

            "film description": "film_description",

            "treatment in": "treatment_in",

            "treatment out": "treatment_out",

            "pallet number": "pallet_number",
            "pallet no": "pallet_number",

            "thickness": "thickness",
            "thickness uom": "thickness_uom",

            "width": "width",
            "width uom": "width_uom",

            "weight": "weight",
            "weight uom": "weight_uom",

            "quantity": "quantity",
            "quantity uom": "quantity_uom",

            "length": "length",
            "length uom": "length_uom",

            "received date": "received_date",

            "aging": "aging",

            "core id": "core_id",
            "core": "core_id",

            "no. of joint": "no_of_joint",
            "no of joint": "no_of_joint",
            "no. of joints": "no_of_joint",

            "lot number": "lot_number",
            "lot no": "lot_number",

            "film": "film",
        }

        return mapping.get(
            value,
            value.replace(" ", "_"),
        )

    def _get_row_value(
        self,
        headers,
        row,
        field_name,
    ):

        try:
            index = headers.index(
                field_name
            )

        except ValueError:
            return None

        if index >= len(row):
            return None

        return row[index]