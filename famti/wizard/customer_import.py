from odoo import models, fields, _
from odoo.exceptions import UserError, ValidationError

import base64
import io
import logging
import re

from openpyxl import load_workbook


_logger = logging.getLogger(__name__)


class CustomerImportWizard(models.TransientModel):
    _name = "customer.import.wizard"
    _description = "Import Customers from Excel"

    file = fields.Binary(
        string="Excel File",
        required=True,
    )

    filename = fields.Char(
        string="Filename",
    )

    # ============================================================
    # BASIC HELPERS
    # ============================================================

    def _clean(self, value):
        """
        Convert Excel value to clean string.

        None -> ""
        Other values -> stripped string
        """
        if value is None:
            return ""

        return str(value).strip()

    def _normalize_name(self, value):
        """
        Normalize name for matching.

        Example:

            Acorn Packaging
            acorn packaging
            ACORN PACKAGING

        are treated as the same name.
        """

        value = self._clean(value)

        value = re.sub(
            r"\s+",
            " ",
            value,
        )

        return value.lower()

    def _normalize_email(self, value):
        """
        Normalize email for matching.
        """

        return self._clean(value).lower()

    # ============================================================
    # SUPPLIER BOOLEAN
    # ============================================================

    def _parse_supplier(
        self,
        value,
        row_number,
        errors,
    ):
        """
        Convert Excel Supplier value to Boolean.

        TRUE:
            Yes
            YES
            yes
            Y
            y
            True
            TRUE
            1

        FALSE:
            No
            NO
            no
            N
            n
            False
            FALSE
            0

        Blank:
            False

        Anything else:
            Validation error.
        """

        value = self._clean(value)

        # Blank supplier means False.
        if not value:
            return False

        normalized = value.lower()

        true_values = {
            "yes",
            "y",
            "true",
            "1",
            "TRUE",
            "True"
        }

        false_values = {
            "no",
            "n",
            "false",
            "0",
            "False",
            "FALSE"
        }

        if normalized in true_values:
            return True

        if normalized in false_values:
            return False

        errors.append(
            _(
                "Row %s: Supplier must be Yes or No. "
                "Received '%s'."
            )
            % (
                row_number,
                value,
            )
        )

        return False

    # ============================================================
    # COUNTRY
    # ============================================================

    def _find_country(
        self,
        country_code,
        row_number,
        errors,
    ):
        """
        Find country by ISO code.

        Example:

            CA -> Canada
            US -> United States
            IN -> India

        Blank is allowed.
        """

        code = self._clean(
            country_code
        ).upper()

        if not code:
            return False

        country = self.env[
            "res.country"
        ].search(
            [
                (
                    "code",
                    "=",
                    code,
                ),
            ],
            limit=1,
        )

        if not country:

            errors.append(
                _(
                    "Row %s: Country code '%s' "
                    "was not found in Odoo."
                )
                % (
                    row_number,
                    code,
                )
            )

            return False

        return country

    # ============================================================
    # STATE
    # ============================================================

    def _find_state(
        self,
        state_code,
        country,
        row_number,
        errors,
    ):
        """
        Find state by state code + country.

        Example:

            ON + CA -> Ontario, Canada

        Blank state is allowed.
        """

        code = self._clean(
            state_code
        ).upper()

        if not code:
            return False

        if not country:

            errors.append(
                _(
                    "Row %s: State code '%s' was provided "
                    "but Country Code is empty."
                )
                % (
                    row_number,
                    code,
                )
            )

            return False

        state = self.env[
            "res.country.state"
        ].search(
            [
                (
                    "code",
                    "=",
                    code,
                ),
                (
                    "country_id",
                    "=",
                    country.id,
                ),
            ],
            limit=1,
        )

        if not state:

            errors.append(
                _(
                    "Row %s: State code '%s' "
                    "was not found for country '%s'."
                )
                % (
                    row_number,
                    code,
                    country.code,
                )
            )

            return False

        return state

    # ============================================================
    # FIND EXISTING CUSTOMER
    # ============================================================

    def _find_customer(
        self,
        customer_name,
        email,
    ):
        """
        Find existing company.

        Priority:

        1. Exact email
        2. Case-insensitive name

        Supplier status is NOT used for matching.

        Therefore:

            Acorn Packaging / Supplier = No

        and later:

            Acorn Packaging / Supplier = Yes

        still refer to the same customer.
        """

        Partner = self.env[
            "res.partner"
        ]

        email = self._normalize_email(
            email
        )

        customer_name = self._clean(
            customer_name
        )

        # --------------------------------------------------------
        # SEARCH BY EMAIL
        # --------------------------------------------------------

        if email:

            partner = Partner.search(
                [
                    (
                        "email",
                        "=ilike",
                        email,
                    ),
                    (
                        "parent_id",
                        "=",
                        False,
                    ),
                    (
                        "company_type",
                        "=",
                        "company",
                    ),
                ],
                limit=1,
            )

            if partner:
                return partner

        # --------------------------------------------------------
        # SEARCH BY NAME
        # --------------------------------------------------------

        if customer_name:

            partners = Partner.search(
                [
                    (
                        "parent_id",
                        "=",
                        False,
                    ),
                    (
                        "company_type",
                        "=",
                        "company",
                    ),
                ]
            )

            normalized_name = (
                self._normalize_name(
                    customer_name
                )
            )

            for partner in partners:

                if (
                    self._normalize_name(
                        partner.name
                    )
                    == normalized_name
                ):
                    return partner

        return False

    # ============================================================
    # FIND EXISTING CONTACT
    # ============================================================

    def _find_contact(
        self,
        company,
        contact_name,
        email,
    ):
        """
        Find contact under customer.

        Priority:

        1. Email
        2. Name
        """

        if not contact_name:
            return False

        Partner = self.env[
            "res.partner"
        ]

        email = self._normalize_email(
            email
        )

        normalized_name = (
            self._normalize_name(
                contact_name
            )
        )

        contacts = Partner.search(
            [
                (
                    "parent_id",
                    "=",
                    company.id,
                ),
                (
                    "type",
                    "=",
                    "contact",
                ),
            ]
        )

        # --------------------------------------------------------
        # EMAIL
        # --------------------------------------------------------

        if email:

            for contact in contacts:

                if (
                    self._normalize_email(
                        contact.email
                    )
                    == email
                ):
                    return contact

        # --------------------------------------------------------
        # NAME
        # --------------------------------------------------------

        for contact in contacts:

            if (
                self._normalize_name(
                    contact.name
                )
                == normalized_name
            ):
                return contact

        return False

    # ============================================================
    # FIND ADDRESS
    # ============================================================

    def _find_address(
        self,
        company,
        address_type,
    ):
        """
        Find first existing address of type.

        delivery -> Shipping
        invoice  -> Billing
        """

        return self.env[
            "res.partner"
        ].search(
            [
                (
                    "parent_id",
                    "=",
                    company.id,
                ),
                (
                    "type",
                    "=",
                    address_type,
                ),
            ],
            order="id asc",
            limit=1,
        )

    # ============================================================
    # ADDRESS VALUES
    # ============================================================

    def _address_values(
        self,
        company,
        address_type,
        street,
        city,
        state,
        zip_code,
        country,
    ):
        """
        Prepare values for Odoo child partner.
        """

        if address_type == "delivery":
            address_name = (
                f"{company.name} - Shipping"
            )
        else:
            address_name = (
                f"{company.name} - Billing"
            )

        return {
            "parent_id": company.id,

            "type": address_type,

            "name": address_name,

            "street": street or False,

            "city": city or False,

            "state_id": (
                state.id
                if state
                else False
            ),

            "zip": zip_code or False,

            "country_id": (
                country.id
                if country
                else False
            ),
        }

    # ============================================================
    # CHECK ADDRESS DATA
    # ============================================================

    def _has_address_data(
        self,
        street,
        city,
        state,
        zip_code,
        country,
    ):
        """
        Return True if at least one address
        field has data.
        """

        return any(
            [
                street,
                city,
                state,
                zip_code,
                country,
            ]
        )

    # ============================================================
    # MAIN IMPORT
    # ============================================================

    def action_import(self):

        self.ensure_one()

        # ========================================================
        # CHECK FILE
        # ========================================================

        if not self.file:

            raise UserError(
                _(
                    "Please select an Excel file."
                )
            )

        # ========================================================
        # READ XLSX
        # ========================================================

        try:

            file_data = base64.b64decode(
                self.file
            )

            workbook = load_workbook(
                filename=io.BytesIO(
                    file_data
                ),
                read_only=True,
                data_only=True,
            )

            sheet = workbook.active

        except Exception as e:

            _logger.exception(
                "Customer Excel import failed."
            )

            raise UserError(
                _(
                    "Unable to read Excel file.\n\n%s"
                )
                % str(e)
            )

        # ========================================================
        # EXPECTED CLIENT TEMPLATE
        # ========================================================

        expected_headers = [

            "customer",

            "contact name",

            "email id",

            "contact number",

            "supplier",

            "shipping street",

            "shipping city",

            "shipping state code",

            "shipping zip",

            "shipping country code",

            "billing street",

            "billing city",

            "billing state code",

            "billing zip",

            "billing country code",
        ]

        # ========================================================
        # READ HEADER
        # ========================================================

        rows = sheet.iter_rows(
            values_only=True
        )

        try:

            header_row = next(rows)

        except StopIteration:

            raise UserError(
                _(
                    "Excel file is empty."
                )
            )

        headers = []

        for value in header_row:

            header = self._clean(
                value
            ).lower()

            headers.append(
                header
            )

        _logger.info(
            "CUSTOMER IMPORT HEADERS: %s",
            headers,
        )

        # ========================================================
        # CHECK MISSING HEADERS
        # ========================================================

        missing_headers = [
            header
            for header in expected_headers
            if header not in headers
        ]

        if missing_headers:

            raise UserError(
                _(
                    "The Excel template is incorrect.\n\n"
                    "Missing columns:\n\n%s\n\n"
                    "Please use the official customer "
                    "import template."
                )
                % "\n".join(
                    missing_headers
                )
            )

        # ========================================================
        # CHECK DUPLICATE HEADERS
        # ========================================================

        duplicate_headers = []

        for header in set(headers):

            if headers.count(header) > 1:

                duplicate_headers.append(
                    header
                )

        if duplicate_headers:

            raise UserError(
                _(
                    "Duplicate Excel columns found:\n\n%s"
                )
                % "\n".join(
                    duplicate_headers
                )
            )

        # ========================================================
        # HEADER INDEX
        # ========================================================

        header_index = {
            header: index
            for index, header
            in enumerate(headers)
        }

        # ========================================================
        # READ ALL EXCEL ROWS
        # ========================================================

        excel_rows = []

        excel_row_number = 1

        for raw_row in rows:

            excel_row_number += 1

            # ----------------------------------------------------
            # Ignore completely empty rows
            # ----------------------------------------------------

            if not any(
                value is not None
                and str(value).strip()
                for value in raw_row
            ):
                continue

            def get(column):

                index = (
                    header_index[column]
                )

                if index >= len(
                    raw_row
                ):
                    return ""

                return raw_row[index]

            excel_rows.append(
                {

                    "row_number":
                        excel_row_number,

                    "customer":
                        self._clean(
                            get("customer")
                        ),

                    "contact_name":
                        self._clean(
                            get("contact name")
                        ),

                    "email":
                        self._clean(
                            get("email id")
                        ),

                    "phone":
                        self._clean(
                            get("contact number")
                        ),

                    "supplier":
                        self._clean(
                            get("supplier")
                        ),

                    "shipping_street":
                        self._clean(
                            get("shipping street")
                        ),

                    "shipping_city":
                        self._clean(
                            get("shipping city")
                        ),

                    "shipping_state":
                        self._clean(
                            get(
                                "shipping state code"
                            )
                        ),

                    "shipping_zip":
                        self._clean(
                            get("shipping zip")
                        ),

                    "shipping_country":
                        self._clean(
                            get(
                                "shipping country code"
                            )
                        ),

                    "billing_street":
                        self._clean(
                            get("billing street")
                        ),

                    "billing_city":
                        self._clean(
                            get("billing city")
                        ),

                    "billing_state":
                        self._clean(
                            get(
                                "billing state code"
                            )
                        ),

                    "billing_zip":
                        self._clean(
                            get("billing zip")
                        ),

                    "billing_country":
                        self._clean(
                            get(
                                "billing country code"
                            )
                        ),
                }
            )

        # ========================================================
        # NO DATA
        # ========================================================

        if not excel_rows:

            raise UserError(
                _(
                    "No customer records were found "
                    "in the Excel file."
                )
            )

        # ========================================================
        # VALIDATION
        #
        # IMPORTANT:
        # Nothing is written to database before
        # this validation is completely successful.
        # ========================================================

        errors = []

        for row in excel_rows:

            row_number = (
                row["row_number"]
            )

            customer_name = (
                row["customer"]
            )

            # ----------------------------------------------------
            # CUSTOMER REQUIRED
            # ----------------------------------------------------

            if not customer_name:

                errors.append(
                    _(
                        "Row %s: Customer is required."
                    )
                    % row_number
                )

                continue

            # ----------------------------------------------------
            # SUPPLIER
            # ----------------------------------------------------

            row["is_supplier"] = (
                self._parse_supplier(
                    row["supplier"],
                    row_number,
                    errors,
                )
            )

            # ----------------------------------------------------
            # SHIPPING COUNTRY
            # ----------------------------------------------------

            shipping_country = (
                self._find_country(
                    row["shipping_country"],
                    row_number,
                    errors,
                )
            )

            # ----------------------------------------------------
            # SHIPPING STATE
            # ----------------------------------------------------

            shipping_state = (
                self._find_state(
                    row["shipping_state"],
                    shipping_country,
                    row_number,
                    errors,
                )
            )

            # ----------------------------------------------------
            # BILLING COUNTRY
            # ----------------------------------------------------

            billing_country = (
                self._find_country(
                    row["billing_country"],
                    row_number,
                    errors,
                )
            )

            # ----------------------------------------------------
            # BILLING STATE
            # ----------------------------------------------------

            billing_state = (
                self._find_state(
                    row["billing_state"],
                    billing_country,
                    row_number,
                    errors,
                )
            )

            # ----------------------------------------------------
            # SAVE RESOLVED RECORDS
            # ----------------------------------------------------

            row[
                "_shipping_country_record"
            ] = shipping_country

            row[
                "_shipping_state_record"
            ] = shipping_state

            row[
                "_billing_country_record"
            ] = billing_country

            row[
                "_billing_state_record"
            ] = billing_state

        # ========================================================
        # STOP IF VALIDATION FAILED
        # ========================================================

        if errors:

            raise ValidationError(
                _(
                    "Customer import validation failed.\n\n"
                    "Nothing was imported.\n\n"
                    "Errors:\n\n%s"
                )
                % "\n".join(
                    errors
                )
            )

        # ========================================================
        # COUNTERS
        # ========================================================

        created_customers = 0
        updated_customers = 0

        created_contacts = 0
        updated_contacts = 0

        created_addresses = 0
        updated_addresses = 0

        # ========================================================
        # PROCESS ROWS
        # ========================================================

        for row in excel_rows:

            row_number = (
                row["row_number"]
            )

            customer_name = (
                row["customer"]
            )

            contact_name = (
                row["contact_name"]
            )

            email = (
                row["email"]
            )

            phone = (
                row["phone"]
            )

            # ====================================================
            # CUSTOMER
            # ====================================================

            customer = (
                self._find_customer(
                    customer_name,
                    email,
                )
            )

            if row["is_supplier"]:
                customer_values = {
                    "name": customer_name,
                    "company_type": "company",
                    "is_company": True,

                    # Your custom field
                    "is_supplier": True,

                    # Odoo's actual vendor/customer classification
                    "supplier_rank": 1,
                    "customer_rank": 0,
                }
            else:
                customer_values = {
                    "name": customer_name,
                    "company_type": "company",
                    "is_company": True,

                    # Your custom field
                    "is_supplier": False,

                    # Odoo's actual vendor/customer classification
                    "supplier_rank": 0,
                    "customer_rank": 1,
                }

            # ----------------------------------------------------
            # Only update email when supplied.
            # ----------------------------------------------------

            if email:

                customer_values[
                    "email"
                ] = email

            # ----------------------------------------------------
            # Only update phone when supplied.
            # ----------------------------------------------------

            if phone:

                customer_values[
                    "phone"
                ] = phone

            # ----------------------------------------------------
            # UPDATE EXISTING CUSTOMER
            # ----------------------------------------------------

            if customer:

                customer.write(
                    customer_values
                )

                updated_customers += 1

                _logger.info(
                    "CUSTOMER IMPORT: "
                    "UPDATED CUSTOMER id=%s name=%s",
                    customer.id,
                    customer.name,
                )

            # ----------------------------------------------------
            # CREATE CUSTOMER
            # ----------------------------------------------------

            else:

                customer = (
                    self.env[
                        "res.partner"
                    ].create(
                        customer_values
                    )
                )

                created_customers += 1

                _logger.info(
                    "CUSTOMER IMPORT: "
                    "CREATED CUSTOMER id=%s name=%s",
                    customer.id,
                    customer.name,
                )

            # ====================================================
            # CONTACT
            # ====================================================

            if contact_name:

                contact = (
                    self._find_contact(
                        customer,
                        contact_name,
                        email,
                    )
                )

                contact_values = {

                    "parent_id":
                        customer.id,

                    "name":
                        contact_name,

                    "type":
                        "contact",
                }

                if email:

                    contact_values[
                        "email"
                    ] = email

                if phone:

                    contact_values[
                        "phone"
                    ] = phone

                # ------------------------------------------------
                # UPDATE CONTACT
                # ------------------------------------------------

                if contact:

                    contact.write(
                        contact_values
                    )

                    updated_contacts += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "UPDATED CONTACT id=%s "
                        "name=%s",
                        contact.id,
                        contact.name,
                    )

                # ------------------------------------------------
                # CREATE CONTACT
                # ------------------------------------------------

                else:

                    contact = (
                        self.env[
                            "res.partner"
                        ].create(
                            contact_values
                        )
                    )

                    created_contacts += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "CREATED CONTACT id=%s "
                        "name=%s",
                        contact.id,
                        contact.name,
                    )

            # ====================================================
            # SHIPPING ADDRESS
            # ====================================================

            shipping_has_data = (
                self._has_address_data(
                    row["shipping_street"],
                    row["shipping_city"],
                    row["shipping_state"],
                    row["shipping_zip"],
                    row["shipping_country"],
                )
            )

            if shipping_has_data:

                shipping = (
                    self._find_address(
                        customer,
                        "delivery",
                    )
                )

                shipping_values = (
                    self._address_values(
                        customer,
                        "delivery",
                        row["shipping_street"],
                        row["shipping_city"],
                        row[
                            "_shipping_state_record"
                        ],
                        row["shipping_zip"],
                        row[
                            "_shipping_country_record"
                        ],
                    )
                )

                # ------------------------------------------------
                # UPDATE SHIPPING
                # ------------------------------------------------

                if shipping:

                    shipping.write(
                        shipping_values
                    )

                    updated_addresses += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "UPDATED SHIPPING ADDRESS "
                        "id=%s customer=%s",
                        shipping.id,
                        customer.name,
                    )

                # ------------------------------------------------
                # CREATE SHIPPING
                # ------------------------------------------------

                else:

                    shipping = (
                        self.env[
                            "res.partner"
                        ].create(
                            shipping_values
                        )
                    )

                    created_addresses += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "CREATED SHIPPING ADDRESS "
                        "id=%s customer=%s",
                        shipping.id,
                        customer.name,
                    )

            # ====================================================
            # BILLING ADDRESS
            # ====================================================

            billing_has_data = (
                self._has_address_data(
                    row["billing_street"],
                    row["billing_city"],
                    row["billing_state"],
                    row["billing_zip"],
                    row["billing_country"],
                )
            )

            if billing_has_data:

                billing = (
                    self._find_address(
                        customer,
                        "invoice",
                    )
                )

                billing_values = (
                    self._address_values(
                        customer,
                        "invoice",
                        row["billing_street"],
                        row["billing_city"],
                        row["_billing_state_record"],
                        row["billing_zip"],
                        row["_billing_country_record"],
                    )
                )

                # ------------------------------------------------
                # UPDATE BILLING
                # ------------------------------------------------

                if billing:

                    billing.write(
                        billing_values
                    )

                    updated_addresses += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "UPDATED BILLING ADDRESS "
                        "id=%s customer=%s",
                        billing.id,
                        customer.name,
                    )

                # ------------------------------------------------
                # CREATE BILLING
                # ------------------------------------------------

                else:

                    billing = (
                        self.env[
                            "res.partner"
                        ].create(
                            billing_values
                        )
                    )

                    created_addresses += 1

                    _logger.info(
                        "CUSTOMER IMPORT: "
                        "CREATED BILLING ADDRESS "
                        "id=%s customer=%s",
                        billing.id,
                        customer.name,
                    )

            _logger.info(
                "CUSTOMER IMPORT: "
                "ROW %s COMPLETE customer=%s",
                row_number,
                customer.name,
            )

        # ========================================================
        # SUCCESS MESSAGE
        # ========================================================

        message = _(
            "Customer import completed successfully.\n\n"
            "Customers created: %s\n"
            "Customers updated: %s\n\n"
            "Contacts created: %s\n"
            "Contacts updated: %s\n\n"
            "Addresses created: %s\n"
            "Addresses updated: %s"
        ) % (
            created_customers,
            updated_customers,
            created_contacts,
            updated_contacts,
            created_addresses,
            updated_addresses,
        )

        return {
            "type":
                "ir.actions.client",

            "tag":
                "display_notification",

            "params": {

                "title":
                    _("Customer Import"),

                "message":
                    message,

                "type":
                    "success",

                "sticky":
                    False,
            },
        }